from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
import unicodedata

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'dev'  # Replace with a secure key in production

db = SQLAlchemy(app)


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(64), unique=False, nullable=False)
    description = db.Column(db.String(256), nullable=False)
    category = db.Column(db.String(128), nullable=True)
    unit = db.Column(db.String(64), nullable=True)
    quantity = db.Column(db.Float, nullable=False)
    unit_value = db.Column(db.Float, nullable=False)
    total_value = db.Column(db.Float, nullable=False)
    assigned_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Item {self.code} - {self.description}>'


@app.before_first_request
def create_tables():
    db.create_all()


@app.route('/')
def index():
    return render_template('index.html')


def _normalize_header(h):
    if not isinstance(h, str):
        return ''
    # remove accents and lowercase
    nk = unicodedata.normalize('NFKD', h)
    return ''.join(c for c in nk if not unicodedata.combining(c)).strip().lower()


def _find_index_for_header(headers, candidates):
    # headers: list of normalized header strings
    for idx, h in enumerate(headers):
        for c in candidates:
            if c in h:
                return idx
    return None


@app.route('/import', methods=['GET', 'POST'])
def import_items():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('No se subió ningún archivo.', 'danger')
            return redirect(url_for('import_items'))

        filename = file.filename or ''
        if not filename.lower().endswith('.xlsx'):
            flash('Formato no soportado. Utilice un archivo .xlsx', 'danger')
            return redirect(url_for('import_items'))

        try:
            wb = load_workbook(file, read_only=True, data_only=True)
        except Exception as e:
            flash(f'Error abriendo el archivo: {e}', 'danger')
            return redirect(url_for('import_items'))

        sheet = wb[wb.sheetnames[0]]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers_orig = next(rows_iter)
        except StopIteration:
            flash('El archivo está vacío.', 'danger')
            return redirect(url_for('import_items'))

        headers = [_normalize_header(h or '') for h in headers_orig]
        # Candidate lists for matching
        col_map = {
            'code': ['codigo', 'cod'],
            'description': ['descripci', 'descripcion', 'descripcion de'],
            'category': ['categoria', 'categor'],
            'unit': ['unidad', 'unidad de medida', 'unidad de'],
            'quantity': ['cantidad', 'cant'],
            'unit_value': ['valor articulo', 'valor articulo (unitario)', 'valor unitario', 'valor unit'],
            'total_value': ['valor total', 'total']
        }

        indices = {}
        for key, candidates in col_map.items():
            idx = _find_index_for_header(headers, candidates)
            indices[key] = idx

        # At minimum we need code, description, quantity and unit_value
        missing_required = []
        for k in ('code', 'description', 'quantity', 'unit_value'):
            if indices.get(k) is None:
                missing_required.append(k)
        if missing_required:
            flash('Columnas obligatorias faltantes en Excel: ' + ', '.join(missing_required), 'danger')
            return redirect(url_for('import_items'))

        imported = 0
        errors = []
        for row_idx, row in enumerate(rows_iter, start=2):
            if not any(row):
                continue
            try:
                code = str(row[indices['code']]).strip() if indices['code'] is not None and row[indices['code']] is not None else ''
                description = str(row[indices['description']]).strip() if indices['description'] is not None and row[indices['description']] is not None else ''
                category = str(row[indices['category']]).strip() if indices['category'] is not None and row[indices['category']] is not None else None
                unit = str(row[indices['unit']]).strip() if indices['unit'] is not None and row[indices['unit']] is not None else None
                quantity_val = row[indices['quantity']] if indices['quantity'] is not None else None
                unit_value_val = row[indices['unit_value']] if indices['unit_value'] is not None else None
                total_value_val = row[indices['total_value']] if indices['total_value'] is not None else None

                if not code or not description:
                    raise ValueError('Código o descripción vacíos')

                try:
                    quantity = float(quantity_val)
                except Exception:
                    raise ValueError('Cantidad no válida')
                try:
                    unit_value = float(unit_value_val)
                except Exception:
                    # If user provided total_value we can compute unit_value
                    if total_value_val is not None and quantity:
                        unit_value = float(total_value_val) / float(quantity)
                    else:
                        raise ValueError('Valor unitario no válido')

                total_value = quantity * unit_value

                item = Item(code=code, description=description, category=category, unit=unit,
                            quantity=quantity, unit_value=unit_value, total_value=total_value)
                db.session.add(item)
                imported += 1
            except Exception as e:
                errors.append(f'Fila {row_idx}: {e}')

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar elementos: {e}', 'danger')
            return redirect(url_for('import_items'))

        if imported > 0:
            flash(f'{imported} filas importadas exitosamente', 'success')
        if errors:
            flash('Se produjeron errores: ' + '; '.join(errors[:5]), 'warning')
        return redirect(url_for('tracking'))

    # GET
    return render_template('import.html')


@app.route('/assign', methods=['GET', 'POST'])
def assign():
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        description = request.form.get('description', '').strip()
        category = request.form.get('category', '').strip()
        unit = request.form.get('unit', '').strip()
        quantity_raw = request.form.get('quantity', '0')
        unit_value_raw = request.form.get('unit_value', '0')

        # Basic validation
        if not code or not description:
            flash('Código y Descripción son campos obligatorios', 'danger')
            return redirect(url_for('assign'))

        try:
            quantity = float(quantity_raw)
            unit_value = float(unit_value_raw)
        except ValueError:
            flash('Cantidad y Valor Unitario deben ser números válidos', 'danger')
            return redirect(url_for('assign'))

        total_value = quantity * unit_value

        item = Item(code=code, description=description, category=category, unit=unit,
                    quantity=quantity, unit_value=unit_value, total_value=total_value)
        db.session.add(item)
        db.session.commit()

        flash('Artículo asignado con éxito', 'success')
        return redirect(url_for('tracking'))

    # GET
    return render_template('assign.html')


@app.route('/tracking')
def tracking():
    items = Item.query.order_by(Item.assigned_at.desc()).all()
    return render_template('tracking.html', items=items)


@app.route('/item/delete/<int:item_id>', methods=['POST'])
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash('Artículo eliminado', 'success')
    return redirect(url_for('tracking'))


@app.route('/item/edit/<int:item_id>', methods=['GET', 'POST'])
def edit_item(item_id):
    item = Item.query.get_or_404(item_id)
    if request.method == 'POST':
        item.code = request.form.get('code', item.code)
        item.description = request.form.get('description', item.description)
        item.category = request.form.get('category', item.category)
        item.unit = request.form.get('unit', item.unit)
        try:
            item.quantity = float(request.form.get('quantity', item.quantity))
            item.unit_value = float(request.form.get('unit_value', item.unit_value))
        except ValueError:
            flash('Cantidad y Valor Unitario deben ser números válidos', 'danger')
            return redirect(url_for('edit_item', item_id=item_id))
        item.total_value = item.quantity * item.unit_value
        db.session.commit()
        flash('Artículo actualizado', 'success')
        return redirect(url_for('tracking'))
    return render_template('edit_item.html', item=item)


if __name__ == '__main__':
    app.run(debug=True)
